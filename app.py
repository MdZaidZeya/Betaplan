import io
import os
import csv
import json
import base64
from datetime import date, datetime
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
from flask import Flask, render_template, request, jsonify, send_file, Response
from openpyxl import load_workbook

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16 MB

# ═══════════════════════════════════════════════════════════════════════════════
#  STORAGE PATHS — two completely separate data dirs, one per dashboard
# ═══════════════════════════════════════════════════════════════════════════════
DATA_DIR   = Path(os.environ.get("DATA_DIR",   "/tmp/betaplan_data"))
BUILD_DIR  = Path(os.environ.get("BUILD_DIR",  "/tmp/buildstate_data"))
for d in (DATA_DIR, BUILD_DIR):
    d.mkdir(parents=True, exist_ok=True)

# Beta Plan paths
BP_SUMMARY = DATA_DIR / "latest_summary.json"
BP_ROWS    = DATA_DIR / "latest_rows.json"
BP_CHART   = DATA_DIR / "latest_chart.png"

# Build State paths
BS_SUMMARY = BUILD_DIR / "latest_summary.json"
BS_ROWS    = BUILD_DIR / "latest_rows.json"   # stored as CSV text
BS_CHART   = BUILD_DIR / "latest_chart.png"

# ═══════════════════════════════════════════════════════════════════════════════
#  BETA PLAN — constants + helpers (unchanged from v1)
# ═══════════════════════════════════════════════════════════════════════════════
WORKSTREAM_ORDER = [
    "Onboarding", "Setup", "Plan", "Prepare", "Execute",
    "LMDP App", "On-Shift", "Cross-temporal", "KTLO",
]
STATUS_COLORS = {
    "Not Started": "#9DBAE5",
    "In Progress":  "#F2C24E",
    "Blocked":      "#E26D6D",
    "Done":         "#7AB87A",
}


def parse_xlsx(file_bytes: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["Inventory"] if "Inventory" in wb.sheetnames else wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    required = {"ID", "Workstream", "Screen", "PctComplete", "Status"}
    missing = required - set(idx.keys())
    if missing:
        raise ValueError(f"Missing columns: {', '.join(sorted(missing))}")
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        rows.append({
            "id":         r[idx["ID"]],
            "workstream": str(r[idx["Workstream"]] or ""),
            "parent":     str(r[idx["Parent"]] or "") if "Parent" in idx else "",
            "screen":     str(r[idx["Screen"]] or ""),
            "owner":      str(r[idx["Owner"]] or "") if "Owner" in idx else "",
            "pct":        int(r[idx["PctComplete"]] or 0),
            "status":     str(r[idx["Status"]] or "Not Started"),
            "lastUpdate": str(r[idx["LastUpdate"]] or "") if "LastUpdate" in idx else "",
            "notes":      str(r[idx["Notes"]] or "") if "Notes" in idx else "",
        })
    return rows


def bp_make_chart(rows: list[dict], snapshot_date: str | None = None) -> bytes:
    snap = snapshot_date or date.today().isoformat()
    grouped = {ws: [] for ws in WORKSTREAM_ORDER}
    for r in rows:
        grouped.setdefault(r["workstream"], []).append(r)
    for ws in grouped:
        grouped[ws].sort(key=lambda r: (r["id"] if isinstance(r["id"], (int, float)) else 999))
    items = []
    for ws in WORKSTREAM_ORDER:
        lanes = grouped.get(ws, [])
        if not lanes:
            continue
        items.append({"kind": "header", "label": ws.upper()})
        for lane in lanes:
            label = lane["screen"]
            if lane["parent"] and lane["parent"] not in ("", ws, lane["workstream"]):
                label = f"  {lane['parent']} > {lane['screen']}"
            items.append({"kind": "lane", "label": label, "pct": lane["pct"],
                          "status": lane["status"], "owner": lane["owner"]})
    n = len(items)
    fig_h = max(8, n * 0.32)
    fig, ax = plt.subplots(figsize=(14, fig_h))
    y_positions = list(range(n))[::-1]
    for y, item in zip(y_positions, items):
        if item["kind"] == "header":
            ax.barh(y, 100, color="#222222", height=0.72, alpha=0.88)
            ax.text(50, y, item["label"], color="white", fontsize=10,
                    weight="bold", ha="center", va="center")
        else:
            color = STATUS_COLORS.get(item["status"], "#CCCCCC")
            ax.barh(y, item["pct"], color=color, height=0.62, edgecolor="#999999", linewidth=0.4)
            ax.barh(y, 100, color="none", edgecolor="#DDDDDD", height=0.62, linewidth=0.4, zorder=0)
            lx = item["pct"] + 1 if item["pct"] < 88 else item["pct"] - 6
            lc = "#333333" if item["pct"] < 88 else "white"
            ax.text(lx, y, f"{item['pct']}%", fontsize=8, va="center", color=lc, weight="bold")
            if item.get("owner"):
                ax.text(102, y, item["owner"], fontsize=7, va="center", color="#888888")
    ax.set_yticks(y_positions)
    ax.set_yticklabels([it["label"] for it in items], fontsize=8.5)
    ax.set_xlim(0, 115)
    ax.set_xlabel("Progress (%)", fontsize=10)
    ax.set_xticks([0, 25, 50, 75, 100])
    ax.set_title(f"Beta Plan 2026 — Progress Snapshot  ({snap})", fontsize=13, weight="bold", pad=14)
    for side in ("top", "right"):
        ax.spines[side].set_visible(False)
    ax.spines["left"].set_color("#CCCCCC")
    ax.spines["bottom"].set_color("#CCCCCC")
    ax.tick_params(colors="#666666")
    ax.set_axisbelow(True)
    ax.grid(axis="x", color="#EEEEEE", linestyle="-", linewidth=0.5)
    legend_handles = [Patch(facecolor=c, edgecolor="#999999", label=s) for s, c in STATUS_COLORS.items()]
    ax.legend(handles=legend_handles, loc="lower right", fontsize=9, frameon=False)
    plt.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=140, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def bp_summarise(rows: list[dict], uploaded_by: str = "") -> dict:
    total = len(rows)
    avg = round(sum(r["pct"] for r in rows) / total) if total else 0
    by_status: dict[str, int] = {}
    for r in rows:
        by_status[r["status"]] = by_status.get(r["status"], 0) + 1
    ws_summary = {}
    for ws in WORKSTREAM_ORDER:
        lanes = [r for r in rows if r["workstream"] == ws]
        if not lanes:
            continue
        ws_avg = round(sum(r["pct"] for r in lanes) / len(lanes))
        ws_summary[ws] = {"avg": ws_avg, "count": len(lanes), "by_status": {}}
        for r in lanes:
            s = r["status"]
            ws_summary[ws]["by_status"][s] = ws_summary[ws]["by_status"].get(s, 0) + 1
    return {
        "total": total, "avg": avg, "by_status": by_status, "workstreams": ws_summary,
        "snapshot_date": date.today().isoformat(),
        "updated_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
        "uploaded_by": uploaded_by,
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  BUILD STATE — constants + helpers (mirrors build_state_chart.py exactly)
# ═══════════════════════════════════════════════════════════════════════════════
JIM_KEYS  = ["Jim_Mockup", "Jim_SP"]
TEAM_KEYS = ["Team_Design", "Team_FEdev", "Team_MW", "Team_FEwiring"]
SIZE_PTS = {"S": 1, "M": 2, "L": 3, "XL": 5}
SUB_PTS  = {"S": 1, "M": 2, "L": 3}

ENGINES = [
    ("Ledger",                  8, 90),
    ("Scheduling brain",        5, 80),
    ("Recommendation engine",   5, 40),
    ("Comms",                   5, 15),
    ("Asset system",            3, 90),
    ("LMDP analytics",          3, 75),
    ("Platform / KTLO",         3, 45),
]
RESERVE_PTS = 7


def _num(v):
    try:
        s = str(v).strip()
        return 0.0 if s in ("", "NA", "N/A") else float(s)
    except (TypeError, ValueError):
        return 0.0


def bs_component_pct(row: dict) -> tuple[float, float]:
    j_vals = [_num(row[k]) for k in JIM_KEYS if str(row.get(k, "")).strip() not in ("NA", "N/A", "")]
    t_vals = [_num(row[k]) for k in TEAM_KEYS if str(row.get(k, "")).strip() not in ("NA", "N/A", "")]
    jim  = sum(j_vals) / len(j_vals) if j_vals else 0.0
    team = sum(t_vals) / len(t_vals) if t_vals else 0.0
    return jim, team


def bs_roll_up(rows: list[dict]) -> dict:
    by_parent: dict[str, list] = {}
    for r in rows:
        if not r.get("Parent Module"):
            continue
        by_parent.setdefault(r["Parent Module"], []).append(r)
    
    out = {}
    for parent, p_rows in by_parent.items():
        by_child = {}
        for r in p_rows:
            by_child.setdefault(r.get("Child Module", ""), []).append(r)
            
        p_wsum = p_jacc = p_tacc = 0.0
        children_out = {}
        
        for child, c_rows in by_child.items():
            c_wsum = c_jacc = c_tacc = 0.0
            child_items = []
            for c in c_rows:
                sz = SIZE_PTS.get(str(c.get("ScreenSize", "M")).strip(), 2)
                sub = SUB_PTS.get(str(c.get("SubSize", "")).strip(), 1)
                w = sz * sub
                j, t = bs_component_pct(c)
                c_wsum += w; c_jacc += w * j; c_tacc += w * t
                child_items.append({
                    "label": c.get("Description", ""),
                    "weight": w,
                    "jim_pct": j,
                    "team_pct": t
                })
            
            p_wsum += c_wsum; p_jacc += c_jacc; p_tacc += c_tacc
            children_out[child] = {
                "weight": c_wsum,
                "jim_pct": c_jacc / c_wsum if c_wsum else 0,
                "team_pct": c_tacc / c_wsum if c_wsum else 0,
                "rows": child_items
            }
            
        out[parent] = {
            "size_pts": p_wsum,
            "jim_pct": p_jacc / p_wsum if p_wsum else 0,
            "team_pct": p_tacc / p_wsum if p_wsum else 0,
            "children": children_out
        }
    return out


def bs_make_chart(rows: list[dict]) -> bytes:
    screens = bs_roll_up(rows)

    # ── build item list ────────────────────────────────────────────────────────
    # Each item: dict with keys: kind, label, indent, jd, td, pct_done
    # kind = "parent" | "child_header" | "row" | "engine_header" | "engine_row" | "gap"
    items = []

    sorted_parents = sorted(
        screens.items(),
        key=lambda kv: -(kv[1]["jim_pct"] * 0.48 + kv[1]["team_pct"] * 0.52),
    )

    for i_par, (mod, p_data) in enumerate(sorted_parents):
        if i_par > 0:
            items.append({"kind": "gap"})          # breathing room between parents
        jd = 0.48 * p_data["jim_pct"]
        td = 0.52 * p_data["team_pct"]
        items.append({"kind": "parent", "label": mod, "jd": jd, "td": td, "pct": jd + td})

        for child, c_data in p_data["children"].items():
            if child:
                c_jd = 0.48 * c_data["jim_pct"]
                c_td = 0.52 * c_data["team_pct"]
                items.append({"kind": "child_header", "label": child,
                               "jd": c_jd, "td": c_td, "pct": c_jd + c_td})
            for r in c_data["rows"]:
                r_jd = 0.48 * r["jim_pct"]
                r_td = 0.52 * r["team_pct"]
                items.append({"kind": "row", "label": r["label"] or "Untitled",
                               "jd": r_jd, "td": r_td, "pct": r_jd + r_td})

    # engines block
    items.append({"kind": "gap"})
    items.append({"kind": "engine_header"})
    for name, pts, pct in ENGINES:
        items.append({"kind": "engine_row", "label": name, "pct": pct})
    items.append({"kind": "engine_row", "label": "Reserve (unscoped)", "pct": 0})

    # ── row heights & y positions ──────────────────────────────────────────────
    ROW_H   = {"parent": 0.52, "child_header": 0.38, "row": 0.45,
               "engine_row": 0.45, "engine_header": 0.0, "gap": 0.0}
    SPACING = {"parent": 0.80, "child_header": 0.60, "row": 0.55,
               "engine_row": 0.55, "engine_header": 0.50, "gap": 0.30}

    n = len(items)
    total_space = sum(SPACING[it["kind"]] for it in items)
    fig_h = max(10, total_space * 0.52)
    fig, ax = plt.subplots(figsize=(15, fig_h))

    # assign y positions (top-down)
    y_pos = []
    y = total_space
    for it in items:
        y -= SPACING[it["kind"]] / 2
        y_pos.append(y)
        y -= SPACING[it["kind"]] / 2

    ax.set_ylim(0, total_space)
    ax.set_xlim(-2, 112)
    ax.invert_yaxis()

    # ── draw ──────────────────────────────────────────────────────────────────
    INDENT_CHILD = 3.5      # x-start for child bars
    INDENT_ROW   = 6.5      # x-start for row bars
    BAR_MAX_CHILD = 96.5    # available bar width for child
    BAR_MAX_ROW   = 93.5

    PARENT_BG  = "#1E293B"   # slate-800
    CHILD_BG   = "#E8F0FE"   # blue-50 tint
    CHILD_LINE = "#93C5FD"   # blue-300
    ENGINE_BG  = "#F1F5F9"   # slate-100

    for yi, it in zip(y_pos, items):
        k = it["kind"]

        if k == "gap":
            continue

        if k == "engine_header":
            ax.axhline(yi, color="#CBD5E1", linewidth=1.5, linestyle="--", zorder=1)
            ax.text(50, yi - 0.08, "── ENGINES ──", fontsize=8.5, color="#94A3B8",
                    ha="center", va="bottom", style="italic")
            continue

        # ── parent: full-width dark band ──────────────────────────────────────
        if k == "parent":
            h = ROW_H["parent"]
            ax.barh(yi, 112, left=-2, height=h + 0.12,
                    color=PARENT_BG, zorder=1)
            # Jim segment
            jd_w = it["jd"] * BAR_MAX_CHILD / 100
            td_w = it["td"] * BAR_MAX_CHILD / 100
            ax.barh(yi, jd_w, left=INDENT_CHILD, height=h * 0.55,
                    color="#60A5FA", zorder=3, label="Jim done (DoR)")
            ax.barh(yi, td_w, left=INDENT_CHILD + jd_w, height=h * 0.55,
                    color="#86EFAC", zorder=3, label="Team done (DoD)")
            remain_w = max(0, BAR_MAX_CHILD - jd_w - td_w)
            ax.barh(yi, remain_w, left=INDENT_CHILD + jd_w + td_w, height=h * 0.55,
                    color=(1, 1, 1, 0.12), zorder=2,
                    edgecolor=(1, 1, 1, 0.2), linewidth=0.4)
            # label on the left
            ax.text(-1.5, yi, it["label"], fontsize=10, weight="bold", color="white",
                    va="center", ha="left", zorder=4)
            # pct on right
            ax.text(103, yi, f"{it['pct']:.0f}%", fontsize=9.5, weight="bold",
                    color="white", va="center", zorder=4)
            continue

        # ── child header: tinted band + left accent line ──────────────────────
        if k == "child_header":
            h = ROW_H["child_header"]
            ax.barh(yi, 114, left=-2, height=h + 0.1, color=CHILD_BG, zorder=1)
            ax.barh(yi, 0.55, left=-2, height=h + 0.1, color=CHILD_LINE, zorder=2)
            ax.text(INDENT_CHILD + 0.5, yi, it["label"].upper(),
                    fontsize=7.5, weight="bold", color="#1D4ED8",
                    va="center", ha="left", zorder=3, style="normal")
            ax.text(103, yi, f"{it['pct']:.0f}%", fontsize=7.5,
                    color="#3B82F6", va="center", zorder=3)
            continue

        # ── row: indented bar ─────────────────────────────────────────────────
        if k == "row":
            h = ROW_H["row"]
            jd_w = it["jd"] * BAR_MAX_ROW / 100
            td_w = it["td"] * BAR_MAX_ROW / 100
            remain_w = max(0, BAR_MAX_ROW - jd_w - td_w)
            ax.barh(yi, jd_w, left=INDENT_ROW, height=h,
                    color="#2563EB", zorder=2, label="Jim done (DoR)")
            ax.barh(yi, td_w, left=INDENT_ROW + jd_w, height=h,
                    color="#7AB87A", zorder=2, label="Team done (DoD)")
            ax.barh(yi, remain_w, left=INDENT_ROW + jd_w + td_w, height=h,
                    color="#E5E7EB", zorder=2, edgecolor="#D1D5DB", linewidth=0.3,
                    label="Remaining")
            ax.text(INDENT_ROW - 0.3, yi, it["label"], fontsize=8, color="#374151",
                    va="center", ha="right")
            ax.text(103, yi, f"{it['pct']:.0f}%", fontsize=7.5,
                    color="#6B7280", va="center")
            continue

        # ── engine row ────────────────────────────────────────────────────────
        if k == "engine_row":
            h = ROW_H["engine_row"]
            pct = it.get("pct", 0)
            ax.barh(yi, pct * BAR_MAX_CHILD / 100, left=INDENT_CHILD, height=h,
                    color="#94A3B8", zorder=2, alpha=0.8)
            ax.barh(yi, max(0, BAR_MAX_CHILD - pct * BAR_MAX_CHILD / 100),
                    left=INDENT_CHILD + pct * BAR_MAX_CHILD / 100, height=h,
                    color="#E5E7EB", zorder=2, edgecolor="#D1D5DB", linewidth=0.3)
            ax.text(INDENT_CHILD - 0.5, yi, it["label"], fontsize=8.5,
                    color="#475569", va="center", ha="right", style="italic")
            ax.text(103, yi, f"{pct:.0f}%", fontsize=8, color="#64748B", va="center")

    # ── axes & chrome ─────────────────────────────────────────────────────────
    ax.set_yticks([])
    ax.set_xticks([0, 25, 50, 75, 100])
    ax.set_xticklabels(["0%", "25%", "50%", "75%", "100%"], fontsize=9, color="#9CA3AF")
    ax.set_xlabel("")
    ax.tick_params(left=False, bottom=True, colors="#D1D5DB")
    for side in ("top", "right", "left"):
        ax.spines[side].set_visible(False)
    ax.spines["bottom"].set_color("#E5E7EB")
    ax.set_axisbelow(True)
    ax.grid(axis="x", color="#F3F4F6", linewidth=0.8, zorder=0)

    # ── x-tick lines at INDENT_CHILD origin ──────────────────────────────────
    for xv in [25, 50, 75, 100]:
        ax.axvline(INDENT_CHILD + xv * BAR_MAX_CHILD / 100,
                   color="#E5E7EB", linewidth=0.5, zorder=0)

    # ── title ─────────────────────────────────────────────────────────────────
    tot_pts = sum(p["size_pts"] for p in screens.values()) + sum(e[1] for e in ENGINES) + RESERVE_PTS
    jd_sum  = sum(p["size_pts"] * 0.48 * p["jim_pct"] / 100 for p in screens.values())
    td_sum  = sum(p["size_pts"] * 0.52 * p["team_pct"] / 100 for p in screens.values())
    done_pts = jd_sum + td_sum + sum(e[1] * e[2] / 100 for e in ENGINES)
    overall_pct = done_pts / tot_pts * 100 if tot_pts else 0
    snap = date.today().isoformat()

    ax.set_title(
        f"Build State  ·  {done_pts:.1f} / {tot_pts:.0f} pts  ({overall_pct:.0f}% complete)  ·  {snap}",
        fontsize=13, weight="bold", color="#1E293B", pad=16, loc="left",
    )

    # ── legend ────────────────────────────────────────────────────────────────
    from matplotlib.patches import Patch as _Patch
    legend_items = [
        _Patch(facecolor="#2563EB", label="Jim done (DoR)"),
        _Patch(facecolor="#7AB87A", label="Team done (DoD)"),
        _Patch(facecolor="#E5E7EB", edgecolor="#D1D5DB", label="Remaining"),
    ]
    ax.legend(handles=legend_items, loc="lower right", frameon=True,
              fontsize=9, framealpha=0.9, edgecolor="#E5E7EB",
              facecolor="white", bbox_to_anchor=(1, 0))

    plt.tight_layout(pad=1.5)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def bs_summarise(rows: list[dict], uploaded_by: str = "") -> dict:
    screens = bs_roll_up(rows)
    tot  = sum(p["size_pts"] for p in screens.values()) + sum(e[1] for e in ENGINES) + RESERVE_PTS
    jim_done  = sum(p["size_pts"] * 0.48 * p["jim_pct"] / 100 for p in screens.values())
    jim_total = sum(p["size_pts"] * 0.48 for p in screens.values())
    team_done  = sum(p["size_pts"] * 0.52 * p["team_pct"] / 100 for p in screens.values())
    team_total = sum(p["size_pts"] * 0.52 for p in screens.values())
    done = jim_done + team_done + sum(e[1] * e[2] / 100 for e in ENGINES)
    overall_pct = round(done / tot * 100) if tot else 0

    by_module = {
        mod: {
            "size_pts": p["size_pts"],
            "jim_pct":  round(p["jim_pct"]),
            "team_pct": round(p["team_pct"]),
            "overall_pct": round(p["jim_pct"]*0.48 + p["team_pct"]*0.52),
        }
        for mod, p in sorted(screens.items(), key=lambda kv: -(kv[1]["jim_pct"]*0.48 + kv[1]["team_pct"]*0.52))
    }
    return {
        "overall_pct": overall_pct,
        "total_pts":   round(tot, 1),
        "done_pts":    round(done, 1),
        "jim_pct":     round(jim_done / jim_total * 100) if jim_total else 0,
        "team_pct":    round(team_done / team_total * 100) if team_total else 0,
        "module_count": len(screens),
        "component_count": len(rows),
        "by_module": by_module,
        "snapshot_date": date.today().isoformat(),
        "updated_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
        "uploaded_by": uploaded_by,
    }


def bs_parse_csv(file_bytes: bytes) -> list[dict]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    required = {"Parent Module", "Child Module", "Description", "ScreenSize", "SubSize",
                 "Jim_Mockup", "Jim_SP", "Team_Design", "Team_FEdev", "Team_MW", "Team_FEwiring"}
    rows = list(reader)
    if not rows:
        raise ValueError("CSV is empty")
    missing = required - set(rows[0].keys())
    if missing:
        raise ValueError(f"Missing CSV columns: {', '.join(sorted(missing))}")
    return rows


def bs_parse_xlsx(file_bytes: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["Inventory"] if "Inventory" in wb.sheetnames else wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    return [dict(zip(headers, [str(c.value) if c.value is not None else "" for c in row]))
            for row in ws.iter_rows(min_row=2)
            if any(c.value for c in row)]


# ═══════════════════════════════════════════════════════════════════════════════
#  ROUTES — BETA PLAN  (all unchanged, prefix /)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    has_data = BP_SUMMARY.exists()
    summary = chart_b64 = None
    if has_data:
        data = json.loads(BP_SUMMARY.read_text())
        chart_b64 = data.pop("chart_b64", None)
        summary = data
    return render_template("index.html", has_data=has_data, summary=summary, chart_b64=chart_b64)


@app.route("/latest-chart")
def latest_chart():
    if BP_CHART.exists():
        return send_file(BP_CHART, mimetype="image/png", download_name="Progress_Chart_latest.png")
    if not BP_ROWS.exists():
        return Response("No chart yet.", status=404, mimetype="text/plain")
    rows = json.loads(BP_ROWS.read_text())
    png  = bp_make_chart(rows)
    BP_CHART.write_bytes(png)
    return send_file(io.BytesIO(png), mimetype="image/png", download_name="Progress_Chart_latest.png")


@app.route("/latest-summary")
def latest_summary():
    if not BP_SUMMARY.exists():
        return jsonify({"error": "No data yet"}), 404
    data = json.loads(BP_SUMMARY.read_text())
    data.pop("chart_b64", None)
    return jsonify(data)


@app.route("/upload", methods=["POST"])
def upload():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Please upload an .xlsx or .xls file"}), 400
    uploaded_by = request.form.get("uploaded_by", "").strip() or "Rishab"
    try:
        rows = parse_xlsx(f.read())
    except ValueError as e:
        return jsonify({"error": str(e)}), 422
    except Exception as e:
        return jsonify({"error": f"Could not read file: {e}"}), 500
    try:
        png = bp_make_chart(rows)
    except Exception as e:
        return jsonify({"error": f"Chart generation failed: {e}"}), 500
    summary = bp_summarise(rows, uploaded_by)
    chart_b64 = base64.b64encode(png).decode()
    BP_SUMMARY.write_text(json.dumps({**summary, "chart_b64": chart_b64}))
    BP_ROWS.write_text(json.dumps(rows))
    BP_CHART.write_bytes(png)
    return jsonify({"ok": True, "summary": summary, "rows": rows})


@app.route("/chart", methods=["POST"])
def chart():
    body = request.get_json(silent=True) or {}
    rows = body.get("rows")
    if not rows:
        return jsonify({"error": "No data"}), 400
    try:
        png = bp_make_chart(rows)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return send_file(io.BytesIO(png), mimetype="image/png",
                     download_name=f"Progress_Chart_{date.today().isoformat()}.png")


# ═══════════════════════════════════════════════════════════════════════════════
#  ROUTES — BUILD STATE  (all under /build prefix)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/build")
def build_index():
    has_data = BS_SUMMARY.exists()
    summary = chart_b64 = None
    if has_data:
        data = json.loads(BS_SUMMARY.read_text())
        chart_b64 = data.pop("chart_b64", None)
        summary = data
    return render_template("build.html", has_data=has_data, summary=summary, chart_b64=chart_b64)


@app.route("/build/latest-chart")
def build_latest_chart():
    if BS_CHART.exists():
        return send_file(BS_CHART, mimetype="image/png", download_name="Build_State_Chart_latest.png")
    if not BS_ROWS.exists():
        return Response("No chart yet.", status=404, mimetype="text/plain")
    rows = json.loads(BS_ROWS.read_text())
    png  = bs_make_chart(rows)
    BS_CHART.write_bytes(png)
    return send_file(io.BytesIO(png), mimetype="image/png", download_name="Build_State_Chart_latest.png")


@app.route("/build/latest-summary")
def build_latest_summary():
    if not BS_SUMMARY.exists():
        return jsonify({"error": "No data yet"}), 404
    data = json.loads(BS_SUMMARY.read_text())
    data.pop("chart_b64", None)
    return jsonify(data)


@app.route("/build/upload", methods=["POST"])
def build_upload():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400
    fname = f.filename.lower()
    uploaded_by = request.form.get("uploaded_by", "").strip() or "Rishab"
    try:
        raw = f.read()
        if fname.endswith(".csv"):
            rows = bs_parse_csv(raw)
        elif fname.endswith((".xlsx", ".xls")):
            rows = bs_parse_xlsx(raw)
        else:
            return jsonify({"error": "Upload a .csv or .xlsx file"}), 400
    except ValueError as e:
        return jsonify({"error": str(e)}), 422
    except Exception as e:
        return jsonify({"error": f"Could not read file: {e}"}), 500
    try:
        png = bs_make_chart(rows)
    except Exception as e:
        return jsonify({"error": f"Chart generation failed: {e}"}), 500
    summary   = bs_summarise(rows, uploaded_by)
    chart_b64 = base64.b64encode(png).decode()
    BS_SUMMARY.write_text(json.dumps({**summary, "chart_b64": chart_b64}))
    BS_ROWS.write_text(json.dumps(rows))
    BS_CHART.write_bytes(png)
    return jsonify({"ok": True, "summary": summary})


@app.route("/build/chart", methods=["POST"])
def build_chart():
    body = request.get_json(silent=True) or {}
    rows = body.get("rows")
    if not rows:
        return jsonify({"error": "No data"}), 400
    try:
        png = bs_make_chart(rows)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return send_file(io.BytesIO(png), mimetype="image/png",
                     download_name=f"Build_State_Chart_{date.today().isoformat()}.png")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    app.run(debug=True, host="0.0.0.0", port=port)
